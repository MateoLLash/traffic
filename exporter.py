"""
exporter.py - Módulo de exportación de datos

Este módulo maneja la exportación de resultados a Excel y CSV
con formatos estructurados y visualizaciones.
"""

import os
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

from utils import setup_logger, format_time

# Configurar logger
logger = setup_logger('exporter')


class DataExporter:
    """
    Exportador de datos de análisis de tráfico
    
    Attributes:
        output_dir: Directorio de salida
        video_name: Nombre del video analizado
        timestamp: Timestamp de la exportación
    """
    
    def __init__(self, output_dir: str, video_name: str = "video"):
        """
        Inicializa el exportador
        
        Args:
            output_dir: Directorio donde guardar archivos
            video_name: Nombre del video
        """
        self.output_dir = output_dir
        self.video_name = video_name
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Crear directorio si no existe
        os.makedirs(output_dir, exist_ok=True)
        
        logger.info(f"Exportador inicializado: {output_dir}")
    
    def export_to_excel(self,
                       statistics: Dict,
                       crossings: List[Dict],
                       time_series: Dict = None,
                       video_properties: Dict = None) -> str:
        """
        Exporta datos completos a Excel con múltiples hojas
        
        Args:
            statistics: Estadísticas del análisis
            crossings: Lista de cruces detectados
            time_series: Datos de series temporales
            video_properties: Propiedades del video
            
        Returns:
            Ruta del archivo generado
        """
        filename = f"{self.video_name}_analisis_{self.timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            # Crear writer de Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Hoja 1: Resumen General
                self._write_summary_sheet(writer, statistics, video_properties)
                
                # Hoja 2: Conteos Totales
                self._write_total_counts_sheet(writer, statistics)
                
                # Hoja 3: Conteos por Línea
                self._write_counts_by_line_sheet(writer, statistics)
                
                # Hoja 4: Detalle de Cruces
                self._write_crossings_sheet(writer, crossings)
                
                # Hoja 5: Series Temporales
                if time_series:
                    self._write_time_series_sheet(writer, time_series)
                
                # Hoja 6: Estadísticas por Dirección
                self._write_direction_stats_sheet(writer, statistics)
            
            # Aplicar formato adicional
            self._apply_excel_formatting(filepath)
            
            logger.info(f"Excel generado: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error al generar Excel: {e}")
            raise
    
    def _write_summary_sheet(self, writer, statistics: Dict, video_properties: Dict):
        """Escribe hoja de resumen general"""
        data = {
            'Métrica': [],
            'Valor': []
        }
        
        # Información del video
        if video_properties:
            data['Métrica'].extend([
                '=== INFORMACIÓN DEL VIDEO ===',
                'Resolución',
                'FPS',
                'Duración (segundos)',
                'Total de Frames',
                ''
            ])
            data['Valor'].extend([
                '',
                f"{video_properties.get('width', 0)}x{video_properties.get('height', 0)}",
                f"{video_properties.get('fps', 0):.2f}",
                video_properties.get('duration', 0),
                video_properties.get('frame_count', 0),
                ''
            ])
        
        # Estadísticas generales
        data['Métrica'].extend([
            '=== ESTADÍSTICAS GENERALES ===',
            'Total de Cruces Detectados',
            'Número de Líneas de Conteo',
            'Tiempo Analizado (seg)',
            'Tiempo Analizado (min)',
            ''
        ])
        data['Valor'].extend([
            '',
            statistics.get('total_crossings', 0),
            statistics.get('total_lines', 0),
            f"{statistics.get('elapsed_time', 0):.2f}",
            f"{statistics.get('elapsed_minutes', 0):.2f}",
            ''
        ])
        
        # Conteos por categoría
        data['Métrica'].append('=== CONTEOS POR CATEGORÍA ===')
        data['Valor'].append('')
        
        total_counts = statistics.get('total_counts', {})
        for class_name, count in total_counts.items():
            data['Métrica'].append(class_name)
            data['Valor'].append(count)
        
        # Promedios
        if 'average_per_minute' in statistics:
            data['Métrica'].append('')
            data['Valor'].append('')
            data['Métrica'].append('=== PROMEDIOS POR MINUTO ===')
            data['Valor'].append('')
            
            for class_name, avg in statistics['average_per_minute'].items():
                data['Métrica'].append(f"{class_name} (prom/min)")
                data['Valor'].append(f"{avg:.2f}")
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Resumen', index=False)
    
    def _write_total_counts_sheet(self, writer, statistics: Dict):
        """Escribe hoja de conteos totales"""
        total_counts = statistics.get('total_counts', {})
        
        data = {
            'Categoría': list(total_counts.keys()),
            'Conteo Total': list(total_counts.values())
        }
        
        # Agregar porcentajes
        total = sum(total_counts.values())
        if total > 0:
            data['Porcentaje (%)'] = [
                f"{(count/total)*100:.1f}" for count in total_counts.values()
            ]
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Conteos Totales', index=False)
    
    def _write_counts_by_line_sheet(self, writer, statistics: Dict):
        """Escribe hoja de conteos por línea"""
        counts_by_line = statistics.get('counts_by_line', [])
        
        rows = []
        for line_info in counts_by_line:
            line_name = line_info['line_name']
            total_count = line_info['total_count']
            counts = line_info['counts_by_class']
            
            for class_name, directions in counts.items():
                rows.append({
                    'Línea': line_name,
                    'Categoría': class_name,
                    'Total': directions['total'],
                    'Arriba → Abajo': directions.get('up_to_down', 0),
                    'Abajo → Arriba': directions.get('down_to_up', 0),
                    'Izquierda → Derecha': directions.get('left_to_right', 0),
                    'Derecha → Izquierda': directions.get('right_to_left', 0)
                })
        
        df = pd.DataFrame(rows)
        if not df.empty:
            df.to_excel(writer, sheet_name='Conteos por Línea', index=False)
    
    def _write_crossings_sheet(self, writer, crossings: List[Dict]):
        """Escribe hoja de detalle de cruces"""
        if not crossings:
            # Crear hoja vacía
            df = pd.DataFrame(columns=['Timestamp', 'Track ID', 'Categoría', 
                                      'Línea', 'Dirección'])
            df.to_excel(writer, sheet_name='Detalle de Cruces', index=False)
            return
        
        rows = []
        for crossing in crossings:
            rows.append({
                'Timestamp (seg)': f"{crossing['timestamp']:.2f}",
                'Timestamp': format_time(crossing['timestamp']),
                'Track ID': crossing['track_id'],
                'Categoría': crossing['class_spanish'],
                'Línea': crossing['line_name'],
                'Dirección': self._format_direction(crossing['direction']),
                'Posición X': crossing['centroid'][0],
                'Posición Y': crossing['centroid'][1]
            })
        
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='Detalle de Cruces', index=False)
    
    def _write_time_series_sheet(self, writer, time_series: Dict):
        """Escribe hoja de series temporales"""
        intervals = time_series['intervals']
        counts = time_series['counts']
        timestamps = time_series['timestamps']
        
        # Crear DataFrame
        data = {'Intervalo': intervals, 'Timestamp (seg)': timestamps}
        data.update(counts)
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Series Temporales', index=False)
    
    def _write_direction_stats_sheet(self, writer, statistics: Dict):
        """Escribe hoja de estadísticas por dirección"""
        counts_by_line = statistics.get('counts_by_line', [])
        
        # Agregar conteos por dirección
        direction_totals = {
            'up_to_down': 0,
            'down_to_up': 0,
            'left_to_right': 0,
            'right_to_left': 0
        }
        
        for line_info in counts_by_line:
            counts = line_info['counts_by_class']
            for class_counts in counts.values():
                for direction in direction_totals.keys():
                    direction_totals[direction] += class_counts.get(direction, 0)
        
        data = {
            'Dirección': [
                'Arriba → Abajo',
                'Abajo → Arriba',
                'Izquierda → Derecha',
                'Derecha → Izquierda'
            ],
            'Conteo Total': [
                direction_totals['up_to_down'],
                direction_totals['down_to_up'],
                direction_totals['left_to_right'],
                direction_totals['right_to_left']
            ]
        }
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Estadísticas Dirección', index=False)
    
    def _apply_excel_formatting(self, filepath: str):
        """Aplica formato a las hojas de Excel"""
        try:
            wb = openpyxl.load_workbook(filepath)
            
            # Estilo para encabezados
            header_fill = PatternFill(start_color="366092", end_color="366092", 
                                     fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Aplicar formato a cada hoja
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Formato de encabezados
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Ajustar ancho de columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
            logger.info("Formato aplicado al Excel")
            
        except Exception as e:
            logger.warning(f"No se pudo aplicar formato al Excel: {e}")
    
    def _format_direction(self, direction: str) -> str:
        """Formatea el nombre de la dirección"""
        direction_map = {
            'up_to_down': 'Arriba → Abajo',
            'down_to_up': 'Abajo → Arriba',
            'left_to_right': 'Izquierda → Derecha',
            'right_to_left': 'Derecha → Izquierda'
        }
        return direction_map.get(direction, direction)
    
    def export_to_csv(self, crossings: List[Dict]) -> str:
        """
        Exporta cruces a CSV (datos crudos)
        
        Args:
            crossings: Lista de cruces
            
        Returns:
            Ruta del archivo generado
        """
        filename = f"{self.video_name}_cruces_raw_{self.timestamp}.csv"
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            if not crossings:
                # Crear CSV vacío
                df = pd.DataFrame(columns=['timestamp', 'track_id', 'class', 
                                          'line', 'direction', 'x', 'y'])
                df.to_csv(filepath, index=False, encoding='utf-8-sig')
                return filepath
            
            rows = []
            for crossing in crossings:
                rows.append({
                    'timestamp_seconds': crossing['timestamp'],
                    'timestamp_formatted': format_time(crossing['timestamp']),
                    'track_id': crossing['track_id'],
                    'class_english': crossing['class_name'],
                    'class_spanish': crossing['class_spanish'],
                    'line_id': crossing['line_id'],
                    'line_name': crossing['line_name'],
                    'direction': crossing['direction'],
                    'position_x': crossing['centroid'][0],
                    'position_y': crossing['centroid'][1]
                })
            
            df = pd.DataFrame(rows)
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            
            logger.info(f"CSV generado: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error al generar CSV: {e}")
            raise
    
    def export_summary_csv(self, statistics: Dict) -> str:
        """
        Exporta resumen de estadísticas a CSV
        
        Args:
            statistics: Estadísticas del análisis
            
        Returns:
            Ruta del archivo generado
        """
        filename = f"{self.video_name}_resumen_{self.timestamp}.csv"
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            total_counts = statistics.get('total_counts', {})
            avg_per_minute = statistics.get('average_per_minute', {})
            
            rows = []
            for class_name in total_counts.keys():
                rows.append({
                    'categoria': class_name,
                    'conteo_total': total_counts[class_name],
                    'promedio_por_minuto': avg_per_minute.get(class_name, 0)
                })
            
            df = pd.DataFrame(rows)
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            
            logger.info(f"CSV de resumen generado: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error al generar CSV de resumen: {e}")
            raise
    
    def export_time_series_csv(self, time_series: Dict) -> str:
        """
        Exporta series temporales a CSV
        
        Args:
            time_series: Datos de series temporales
            
        Returns:
            Ruta del archivo generado
        """
        filename = f"{self.video_name}_series_temporales_{self.timestamp}.csv"
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            data = {
                'intervalo': time_series['intervals'],
                'timestamp_seconds': time_series['timestamps']
            }
            data.update(time_series['counts'])
            
            df = pd.DataFrame(data)
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            
            logger.info(f"CSV de series temporales generado: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error al generar CSV de series temporales: {e}")
            raise
    
    def export_all(self,
                   statistics: Dict,
                   crossings: List[Dict],
                   time_series: Dict = None,
                   video_properties: Dict = None) -> Dict[str, str]:
        """
        Exporta todos los formatos disponibles
        
        Args:
            statistics: Estadísticas del análisis
            crossings: Lista de cruces
            time_series: Datos de series temporales
            video_properties: Propiedades del video
            
        Returns:
            Diccionario con rutas de todos los archivos generados
        """
        files = {}
        
        try:
            # Excel completo
            files['excel'] = self.export_to_excel(
                statistics, crossings, time_series, video_properties
            )
            
            # CSV de cruces crudos
            files['csv_crossings'] = self.export_to_csv(crossings)
            
            # CSV de resumen
            files['csv_summary'] = self.export_summary_csv(statistics)
            
            # CSV de series temporales
            if time_series:
                files['csv_timeseries'] = self.export_time_series_csv(time_series)
            
            logger.info(f"Exportación completa: {len(files)} archivos generados")
            return files
            
        except Exception as e:
            logger.error(f"Error en exportación completa: {e}")
            raise


if __name__ == "__main__":
    # Pruebas del exportador
    print("=== Pruebas del Exporter ===")
    
    # Crear exportador
    exporter = DataExporter(output_dir="test_exports", video_name="test_video")
    
    # Datos de prueba
    test_stats = {
        'total_crossings': 100,
        'total_lines': 2,
        'elapsed_time': 300,
        'elapsed_minutes': 5,
        'total_counts': {'Auto': 50, 'Peatón': 30, 'Moto': 20},
        'average_per_minute': {'Auto': 10, 'Peatón': 6, 'Moto': 4},
        'counts_by_line': []
    }
    
    test_crossings = []
    
    print("✓ Módulo exporter funcionando correctamente")
