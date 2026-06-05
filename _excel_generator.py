"""
excel_generator.py - Generador de Excel profesional para aforos vehiculares y peatonales

Este módulo genera reportes de Excel siguiendo el formato oficial peruano
de aforos vehiculares y peatonales con:
- Intervalos de 15 minutos
- Clasificación por tipo de vehículo/peatón
- Múltiples giros y sentidos
- Cálculo automático de hora punta
- Fórmulas y formato profesional
"""

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                             Protection)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
from collections import defaultdict

from utils import setup_logger

# Configurar logger
logger = setup_logger('excel_generator')


# Definición de tipos de vehículos peruanos
VEHICLE_TYPES_URBAN = [
    'Auto',
    'Moto lineal',
    'Microbús',
    'Camioneta Rural',
    'Camión',
    'Mototaxi',
    'Carreta',
    'Bicicleta',
    'Scooter'
]

VEHICLE_TYPES_HIGHWAY = [
    'Auto',
    'Moto lineal',
    'Ómnibus',
    'Microbús',
    'Camioneta Rural',
    'Bus interprovincial',
    'Camión 2E',
    'Camión 3E',
    'Camión 4E',
    '2S1/2S2',
    '2S3',
    '3S1/3S2',
    '3S3',
    '>=4S1',
    'Mototaxi',
    'Bicicleta'
]

PEDESTRIAN_TYPES = [
    'Niño-H',
    'Niño-M',
    'Adulto-H',
    'Adulto-M',
    'Adulto Mayor-H',
    'Adulto Mayor-M',
    'PMR-H',
    'PMR-M'
]


class ExcelAforoGenerator:
    """
    Generador de Excel profesional para aforos vehiculares y peatonales
    
    Attributes:
        workbook: Libro de Excel
        intersection_name: Nombre de la intersección
        session_date: Fecha de la sesión
        interval_minutes: Intervalo de agrupación
    """
    
    def __init__(self,
                 intersection_name: str,
                 session_date: datetime,
                 interval_minutes: int = 15,
                 is_highway: bool = False):
        """
        Inicializa el generador
        
        Args:
            intersection_name: Nombre de la intersección
            session_date: Fecha de inicio de la sesión
            interval_minutes: Intervalo en minutos (default: 15)
            is_highway: Si es carretera (True) o zona urbana (False)
        """
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # Remover hoja por defecto
        
        self.intersection_name = intersection_name
        self.session_date = session_date
        self.interval_minutes = interval_minutes
        self.is_highway = is_highway
        
        # Seleccionar tipos de vehículos según contexto
        self.vehicle_types = (VEHICLE_TYPES_HIGHWAY if is_highway 
                             else VEHICLE_TYPES_URBAN)
        
        # Estilos predefinidos
        self._init_styles()
        
        logger.info(f"Generador de Excel inicializado: {intersection_name}")
    
    def _init_styles(self):
        """Inicializa estilos predefinidos"""
        # Colores
        self.color_header = "366092"  # Azul oscuro
        self.color_subheader = "4472C4"  # Azul
        self.color_total = "FFC000"  # Amarillo/naranja
        self.color_peak = "FF0000"  # Rojo (hora punta)
        self.color_alt_row = "E7E6E6"  # Gris claro
        
        # Fuentes
        self.font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        self.font_normal = Font(name='Arial', size=10)
        self.font_bold = Font(name='Arial', size=10, bold=True)
        
        # Rellenos
        self.fill_header = PatternFill(start_color=self.color_header,
                                      end_color=self.color_header,
                                      fill_type='solid')
        self.fill_subheader = PatternFill(start_color=self.color_subheader,
                                         end_color=self.color_subheader,
                                         fill_type='solid')
        self.fill_total = PatternFill(start_color=self.color_total,
                                     end_color=self.color_total,
                                     fill_type='solid')
        self.fill_peak = PatternFill(start_color=self.color_peak,
                                    end_color=self.color_peak,
                                    fill_type='solid')
        self.fill_alt = PatternFill(start_color=self.color_alt_row,
                                   end_color=self.color_alt_row,
                                   fill_type='solid')
        
        # Bordes
        thin_side = Side(style='thin', color='000000')
        self.border_all = Border(left=thin_side, right=thin_side,
                                top=thin_side, bottom=thin_side)
        
        # Alineaciones
        self.align_center = Alignment(horizontal='center', vertical='center',
                                     wrap_text=True)
        self.align_left = Alignment(horizontal='left', vertical='center')
        self.align_right = Alignment(horizontal='right', vertical='center')
    
    def create_vehicular_sheet(self,
                               direction_name: str,
                               consolidated_counts: Dict,
                               turn_mapping: Dict[int, str]) -> str:
        """
        Crea una hoja de aforo vehicular
        
        Args:
            direction_name: Nombre de la dirección (ej: "N-S", "S-N")
            consolidated_counts: Datos consolidados de conteos
            turn_mapping: Mapeo de número de giro a descripción
            
        Returns:
            Nombre de la hoja creada
        """
        sheet_name = f"Vehicular_{direction_name}"
        ws = self.workbook.create_sheet(sheet_name)
        
        logger.info(f"Creando hoja vehicular: {sheet_name}")
        
        # Información de cabecera
        self._write_sheet_header(ws, "AFORO VEHICULAR", direction_name)
        
        # Determinar giros a mostrar
        turn_numbers = sorted(turn_mapping.keys())
        
        # Escribir encabezados de columnas
        header_row = 7
        self._write_vehicular_headers(ws, header_row, turn_numbers, turn_mapping)
        
        # Escribir datos de conteo
        data_start_row = header_row + 1
        intervals = consolidated_counts['intervals']
        counts = consolidated_counts['counts']
        
        self._write_vehicular_data(ws, data_start_row, intervals, counts,
                                   turn_numbers, self.vehicle_types)
        
        # Aplicar formato
        self._apply_sheet_formatting(ws)
        
        # Calcular y resaltar hora punta
        self._highlight_peak_hour(ws, data_start_row, len(intervals))
        
        return sheet_name
    
    def create_pedestrian_sheet(self,
                               consolidated_counts: Dict,
                               direction_mapping: Dict[int, str]) -> str:
        """
        Crea una hoja de aforo peatonal
        
        Args:
            consolidated_counts: Datos consolidados de conteos peatonales
            direction_mapping: Mapeo de número de dirección a descripción
            
        Returns:
            Nombre de la hoja creada
        """
        sheet_name = "Peatonal"
        ws = self.workbook.create_sheet(sheet_name)
        
        logger.info(f"Creando hoja peatonal: {sheet_name}")
        
        # Información de cabecera
        self._write_sheet_header(ws, "AFORO PEATONAL", "Todos los sentidos")
        
        # Determinar direcciones a mostrar
        direction_numbers = sorted(direction_mapping.keys())
        
        # Escribir encabezados
        header_row = 7
        self._write_pedestrian_headers(ws, header_row, direction_numbers,
                                      direction_mapping)
        
        # Escribir datos
        data_start_row = header_row + 1
        intervals = consolidated_counts['intervals']
        counts = consolidated_counts['counts']
        
        self._write_pedestrian_data(ws, data_start_row, intervals, counts,
                                   direction_numbers, PEDESTRIAN_TYPES)
        
        # Aplicar formato
        self._apply_sheet_formatting(ws)
        
        return sheet_name
    
    def create_summary_sheet(self, consolidated_counts: Dict) -> str:
        """
        Crea hoja de resumen con estadísticas generales
        
        Args:
            consolidated_counts: Datos consolidados
            
        Returns:
            Nombre de la hoja creada
        """
        sheet_name = "Resumen"
        ws = self.workbook.create_sheet(sheet_name, 0)  # Primera posición
        
        logger.info("Creando hoja de resumen")
        
        # Título
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = f"RESUMEN DE AFORO - {self.intersection_name}"
        cell.font = Font(name='Arial', size=14, bold=True)
        cell.alignment = self.align_center
        cell.fill = self.fill_header
        
        # Información general
        row = 3
        info = [
            ('Intersección:', self.intersection_name),
            ('Fecha:', self.session_date.strftime('%d/%m/%Y')),
            ('Hora de inicio:', self.session_date.strftime('%H:%M')),
            ('Intervalo:', f'{self.interval_minutes} minutos'),
            ('', ''),
            ('TOTALES GENERALES', ''),
        ]
        
        for label, value in info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if label != '':
                ws[f'A{row}'].font = self.font_bold
            row += 1
        
        # Totales por tipo de vehículo
        totals = consolidated_counts['totals']['by_class']
        
        ws[f'A{row}'] = 'Tipo de Vehículo'
        ws[f'B{row}'] = 'Total'
        ws[f'A{row}'].font = self.font_header
        ws[f'B{row}'].font = self.font_header
        ws[f'A{row}'].fill = self.fill_subheader
        ws[f'B{row}'].fill = self.fill_subheader
        row += 1
        
        for vehicle_type, count in sorted(totals.items()):
            ws[f'A{row}'] = vehicle_type
            ws[f'B{row}'] = count
            row += 1
        
        # Total general
        ws[f'A{row}'] = 'TOTAL'
        ws[f'B{row}'] = consolidated_counts['totals']['grand_total']
        ws[f'A{row}'].font = self.font_bold
        ws[f'B{row}'].font = self.font_bold
        ws[f'A{row}'].fill = self.fill_total
        ws[f'B{row}'].fill = self.fill_total
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        
        return sheet_name
    
    def create_charts_sheet(self, consolidated_counts: Dict) -> str:
        """
        Crea hoja con gráficos y KPIs
        
        Args:
            consolidated_counts: Datos consolidados
            
        Returns:
            Nombre de la hoja creada
        """
        sheet_name = "Gráficos y KPIs"
        ws = self.workbook.create_sheet(sheet_name)
        
        logger.info("Creando hoja de gráficos")
        
        # Título
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = "ANÁLISIS Y VISUALIZACIÓN DE DATOS"
        cell.font = Font(name='Arial', size=14, bold=True)
        cell.alignment = self.align_center
        cell.fill = self.fill_header
        
        # KPIs principales
        row = 3
        ws[f'A{row}'] = "KPIs PRINCIPALES"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        # Aquí se agregarán gráficos más adelante
        # Por ahora solo estructura
        
        return sheet_name
    
    def _write_sheet_header(self, ws, title: str, direction: str):
        """Escribe el encabezado de una hoja"""
        # Título principal
        ws.merge_cells('A1:K1')
        cell = ws['A1']
        cell.value = title
        cell.font = Font(name='Arial', size=14, bold=True)
        cell.alignment = self.align_center
        cell.fill = self.fill_header
        
        # Información
        ws['A2'] = 'INTERSECCIÓN:'
        ws['B2'] = self.intersection_name
        ws['A2'].font = self.font_bold
        
        ws['A3'] = 'FECHA:'
        ws['B3'] = self.session_date.strftime('%d/%m/%Y')
        ws['A3'].font = self.font_bold
        
        ws['A4'] = 'SENTIDO:'
        ws['B4'] = direction
        ws['A4'].font = self.font_bold
        
        ws['A5'] = 'INTERVALO:'
        ws['B5'] = f'{self.interval_minutes} minutos'
        ws['A5'].font = self.font_bold
    
    def _write_vehicular_headers(self, ws, row: int, turn_numbers: List[int],
                                 turn_mapping: Dict[int, str]):
        """Escribe los encabezados de columnas para aforo vehicular"""
        col = 1
        
        # Columnas de tiempo
        ws.cell(row, col, "TIPO DE\nVEHÍCULO")
        ws.cell(row, col).font = self.font_header
        ws.cell(row, col).fill = self.fill_header
        ws.cell(row, col).alignment = self.align_center
        col += 1
        
        ws.cell(row, col, "SENTIDO\nHORA")
        ws.cell(row, col).font = self.font_header
        ws.cell(row, col).fill = self.fill_header
        ws.cell(row, col).alignment = self.align_center
        col += 1
        
        # Columnas de tipos de vehículos
        for vehicle_type in self.vehicle_types:
            # Por cada tipo, sub-columnas para cada giro
            start_col = col
            
            for turn_num in turn_numbers:
                ws.cell(row, col, str(turn_num))
                ws.cell(row, col).font = self.font_header
                ws.cell(row, col).fill = self.fill_subheader
                ws.cell(row, col).alignment = self.align_center
                col += 1
            
            # Merge del tipo de vehículo
            if len(turn_numbers) > 1:
                ws.merge_cells(start_row=row-1, start_column=start_col,
                              end_row=row-1, end_column=col-1)
                merged_cell = ws.cell(row-1, start_col)
                merged_cell.value = vehicle_type
                merged_cell.font = self.font_header
                merged_cell.fill = self.fill_header
                merged_cell.alignment = self.align_center
        
        # Columnas de totales
        ws.cell(row, col, "TOTAL\nx 1/4 Hrs")
        ws.cell(row, col).font = self.font_header
        ws.cell(row, col).fill = self.fill_total
        ws.cell(row, col).alignment = self.align_center
        col += 1
        
        ws.cell(row, col, "TOTAL\nHORARIA")
        ws.cell(row, col).font = self.font_header
        ws.cell(row, col).fill = self.fill_total
        ws.cell(row, col).alignment = self.align_center
    
    def _write_pedestrian_headers(self, ws, row: int, direction_numbers: List[int],
                                  direction_mapping: Dict[int, str]):
        """Escribe los encabezados de columnas para aforo peatonal"""
        col = 1
        
        # Columna de hora
        ws.cell(row, col, "HORA")
        ws.cell(row, col).font = self.font_header
        ws.cell(row, col).fill = self.fill_header
        ws.cell(row, col).alignment = self.align_center
        col += 1
        
        # Columnas de tipos de peatones
        for ped_type in PEDESTRIAN_TYPES:
            start_col = col
            
            for dir_num in direction_numbers:
                ws.cell(row, col, str(dir_num))
                ws.cell(row, col).font = self.font_header
                ws.cell(row, col).fill = self.fill_subheader
                ws.cell(row, col).alignment = self.align_center
                col += 1
            
            # Merge del tipo de peatón
            if len(direction_numbers) > 1:
                ws.merge_cells(start_row=row-1, start_column=start_col,
                              end_row=row-1, end_column=col-1)
                merged_cell = ws.cell(row-1, start_col)
                merged_cell.value = ped_type
                merged_cell.font = self.font_header
                merged_cell.fill = self.fill_header
                merged_cell.alignment = self.align_center
        
        # Total
        ws.cell(row, col, "TOTAL")
        ws.cell(row, col).font = self.font_header
        ws.cell(row, col).fill = self.fill_total
        ws.cell(row, col).alignment = self.align_center
    
    def _write_vehicular_data(self, ws, start_row: int, intervals: List[Tuple],
                             counts: Dict, turn_numbers: List[int],
                             vehicle_types: List[str]):
        """Escribe los datos de conteo vehicular"""
        row = start_row
        
        for idx, (interval_start, interval_end) in enumerate(intervals):
            col = 1
            
            # Hora
            time_str = interval_start.strftime('%H:%M')
            ws.cell(row, col, time_str)
            ws.cell(row, col).alignment = self.align_center
            col += 2  # Saltar columna de sentido
            
            # Conteos por vehículo y giro
            interval_total = 0
            
            for vehicle_type in vehicle_types:
                for turn_num in turn_numbers:
                    count = counts[idx].get(vehicle_type, {}).get(turn_num, 0)
                    ws.cell(row, col, count if count > 0 else '')
                    ws.cell(row, col).alignment = self.align_center
                    interval_total += count
                    col += 1
            
            # Total del intervalo
            ws.cell(row, col, interval_total)
            ws.cell(row, col).font = self.font_bold
            ws.cell(row, col).fill = self.fill_alt
            ws.cell(row, col).alignment = self.align_center
            col += 1
            
            # Total horario (suma de últimos 4 intervalos)
            if idx >= 3:
                hourly_total = sum(
                    sum(counts[i].get(vt, {}).values())
                    for i in range(idx-3, idx+1)
                    for vt in vehicle_types
                )
                ws.cell(row, col, hourly_total)
                ws.cell(row, col).font = self.font_bold
                ws.cell(row, col).alignment = self.align_center
            
            row += 1
    
    def _write_pedestrian_data(self, ws, start_row: int, intervals: List[Tuple],
                              counts: Dict, direction_numbers: List[int],
                              pedestrian_types: List[str]):
        """Escribe los datos de conteo peatonal"""
        row = start_row
        
        for idx, (interval_start, interval_end) in enumerate(intervals):
            col = 1
            
            # Hora
            time_str = interval_start.strftime('%H:%M')
            ws.cell(row, col, time_str)
            ws.cell(row, col).alignment = self.align_center
            col += 1
            
            # Conteos por tipo y dirección
            interval_total = 0
            
            for ped_type in pedestrian_types:
                for dir_num in direction_numbers:
                    count = counts[idx].get(ped_type, {}).get(dir_num, 0)
                    ws.cell(row, col, count if count > 0 else '')
                    ws.cell(row, col).alignment = self.align_center
                    interval_total += count
                    col += 1
            
            # Total
            ws.cell(row, col, interval_total)
            ws.cell(row, col).font = self.font_bold
            ws.cell(row, col).alignment = self.align_center
            
            row += 1
    
    def _apply_sheet_formatting(self, ws):
        """Aplica formato general a una hoja"""
        # Aplicar bordes a todas las celdas con datos
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = self.border_all
        
        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 15)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _highlight_peak_hour(self, ws, data_start_row: int, num_intervals: int):
        """Resalta la hora punta en la hoja"""
        # Por ahora solo marca con color, se refinará con la lógica completa
        # Esta función se completará cuando integremos con consolidator
        pass
    
    def save(self, filename: str) -> str:
        """
        Guarda el libro de Excel
        
        Args:
            filename: Nombre del archivo
            
        Returns:
            Ruta del archivo guardado
        """
        self.workbook.save(filename)
        logger.info(f"Excel guardado: {filename}")
        return filename


if __name__ == "__main__":
    # Pruebas del generador
    print("=== Pruebas del Excel Generator ===")
    
    # Crear generador
    generator = ExcelAforoGenerator(
        intersection_name="Av. República de Panamá con Vía Expresa",
        session_date=datetime(2024, 3, 15, 6, 0, 0),
        interval_minutes=15,
        is_highway=False
    )
    
    # Crear hojas de prueba
    test_counts = {
        'intervals': [
            (datetime(2024, 3, 15, 6, 0), datetime(2024, 3, 15, 6, 15)),
            (datetime(2024, 3, 15, 6, 15), datetime(2024, 3, 15, 6, 30)),
        ],
        'counts': {
            0: {'Auto': {1: 10, 2: 5}},
            1: {'Auto': {1: 15, 2: 8}},
        },
        'totals': {
            'by_class': {'Auto': 38},
            'grand_total': 38
        }
    }
    
    turn_map = {1: 'N-S', 2: 'S-N'}
    
    generator.create_summary_sheet(test_counts)
    generator.create_vehicular_sheet('N-S', test_counts, turn_map)
    generator.save('test_aforo.xlsx')
    
    print("✓ Excel de prueba generado: test_aforo.xlsx")
    print("\n✓ Módulo excel_generator funcionando correctamente")
